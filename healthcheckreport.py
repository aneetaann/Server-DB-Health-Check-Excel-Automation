#!/usr/bin/python3

import paramiko
import config
import cx_Oracle
cx_Oracle.init_oracle_client(lib_dir=r"")
import itertools
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors, Color, Fill, Border, Side, Alignment, PatternFill, GradientFill
from datetime import date, timedelta


hosts = ['ABC0005', 'ABC0006', 'ABC0015', 'ABC0016']
cmd = "df -kh /mount_name|tr -s ' ' | cut -d' ' -f5|tail -n1"
cells = ['I11', 'I12', 'I13', 'I14']                                                #excel cells to be edited for server health check output
cell_mappings = dict(zip(hosts, cells))
cell_values = {}

excel_cells = ('F18','H18','I18','J18')                                             #excel cells to be edited for sql query output

#Date Declaration
today = date.today()
yesterday = today - timedelta(days=1)

#dd-month abbr-YYYY
curr_date = today.strftime("%d-%b-%Y")
yest_date = yesterday.strftime("%d-%b-%Y")
print(yest_date, curr_date)


#SERVER HEALTH CHECK - DISK USAGE
for host in hosts:
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, username=config.username, password=config.password)

    stdin, stdout, stderr = ssh.exec_command(cmd)

    if stdout.channel.recv_exit_status() == 0:
        usepercent = stdout.read().decode("utf8")
        cell_values[host] = usepercent
        for key, value in cell_values.items():
            cell_values[key] = value.rstrip()
        print(f'STDOUT {host}: {usepercent}')
    else:
        print(f'STDERR: {stderr.read().decode("utf8")}')


stdin.close()
stdout.close()
stderr.close()
ssh.close()


#SQL QUERY AUTOMATION
try:
    #create connection
    conn = cx_Oracle.connect(user=config.dbuserid, password=config.dbpwd, dsn=config.connectionstring, encoding='UTF-8')
except Exception as err:
    print('Exception occurred while trying to create a connection',err)
else:
    try:
        #create cursor
        cursor = conn.cursor()
        #sql query for execution
        query = """
        select count(1) from wmptn025.BIZDOC where DOCTIMESTAMP > '{YEST_DATE} 12:00:00' and doctimestamp <='{CURR_DATE} 12:00:00'
        UNION ALL
        select count(1) from wmptn025.BIZDOC where UserStatus LIKE 'ERR%' and doctimestamp > '{YEST_DATE} 12:00:00' and doctimestamp <='{CURR_DATE} 12:00:00'
        UNION ALL
        select count(1) from wmptn025.BIZDOC where UserStatus LIKE 'COMPLETE%' and RoutingStatus LIKE 'REPROCESSED%' and doctimestamp >'{YEST_DATE} 12:00:00' and doctimestamp <='{CURR_DATE} 12:00:00'
        UNION ALL
        select count(1) from wmptn025.BIZDOC where UserStatus LIKE 'QUEUED%' or UserStatus like 'SUBMITTED%' and doctimestamp > '{YEST_DATE} 12:00:00' and doctimestamp <='{CURR_DATE} 12:00:00'
        """.format(CURR_DATE=curr_date, YEST_DATE=yest_date)
        #execute query
        cursor.execute(query)
        #fetchall() outputs all the values obtained from the query.
        global values
        values = cursor.fetchall()                                  #output is stored in a variable "values" in the form of a list of tuples
        values = [list(x) for x in values]                          #converting list of tuples into list of lists
        values = list(itertools.chain(*values))                     #flattening list of lists into a single list
        print(values)
    except Exception as err:
        print('Exception while executing query',err)
    else:
        print('SQL Queries Executed')
    finally:
        cursor.close()
finally:
    conn.close()

#convert "values" into a tuple
transactions = tuple(values)

final_values = dict(zip(excel_cells, transactions))


#EXCEL GENERATION

#open excel workbook
workbook = load_workbook(filename=r"")                            #full filepath where the original health check report is placed eg:C:/healthcheck/report.xlsx

#open active sheet
sheet = workbook.active

#dates
report_date = today.strftime("%d/%B/%Y").upper()
curr_weekday = today.strftime("%A")
today_date = today.strftime("%m/%d/%Y")
 
yester_date = yesterday.strftime("%d/%B/%Y").upper()
yest_weekday = yesterday.strftime("%A")
yes_date = yesterday.strftime("%m/%d/%Y")

friday = today - timedelta(days=3)
fri_date = friday.strftime("%d/%B/%Y").upper()
fri_weekday = friday.strftime("%A")

print(report_date, curr_weekday, yester_date, yest_weekday, fri_date, fri_weekday)	

#EDITING THE EXCEL SHEET

#disk usage output in excel
#loop to iterate through keys in dict(cell_mappings) and find the matching keys in dict(cell_values) and save the values in the corresponding excel cells
for hostkey in cell_mappings.keys():
    if hostkey in cell_values.keys():
        sheet[cell_mappings[hostkey]].value = cell_values[hostkey]
    else:
        print('Mapping Error')

#sql query output in excel
#loop to iterate through the dictionary and map it to specified excel cells by checking key-value combination
try:
    for key in final_values:
        sheet[key].value = final_values[key]                                                 
except Exception as err:
    print('Exception while mapping output to excel cells',err)
else:
    print('Excel Updated With SQL values')



#cell B4
if yest_weekday == 'Sunday':
    sheet['B4'] = "Report Date : " + str(curr_weekday) + " - " + str(report_date) + " 09:00 AM EST                  Duration - From " + str(fri_weekday) + " - " + str(fri_date) + " 08:00 AM EST to " + str(curr_weekday) + " - " + str(report_date) + " 08:00 AM EST.     Prepared and reported by Capgemini Integration support team"     
    b = sheet['B4']
    b.font = Font(name='Calibri', size=10, bold=True, color="B30000")
else:
    sheet['B4'] = "Report Date : " + str(curr_weekday) + " - " + str(report_date) + " 09:00 AM EST                  Duration - From " + str(yest_weekday) + " - " + str(yester_date) + " 08:00 AM EST to " + str(curr_weekday) + " - " + str(report_date) + " 08:00 AM EST.     Prepared and reported by Capgemini Integration support team"     
    b = sheet['B4']
    b.font = Font(name='Calibri', size=10, bold=True, color="B30000")

#cell D16
sheet['D16'] = "webMethod -Transaction Volume and status in TN (" + str(yes_date) + " 08:00 AM - " + str(today_date) + " 08:00 AM)"
d = sheet['D16']
d.font = Font(name='Calibri', size=11, bold=True)



#save as a new excel file
workbook.save(filename = r"" + str(curr_date) + ".xlsx")                         #full filepath along with new name of excel sheet eg:C:/healthcheck/report_date.xlsx

